from langchain_ollama import OllamaLLM
import sqlite3
import openpyxl  # Pour générer un fichier Excel
import os  # Pour vérifier l'existence du fichier
from openpyxl.styles import Font
import time

# Charger les résultats attendus depuis le fichier functions.txt
def parse_expected_results(file_path):
    expected_results = {}
    with open(file_path, "r") as file:
        current_algorithm = None
        for line in file:
            line = line.strip()
            if line.endswith(":"):
                # Extraire le nom de l'algorithme (ex : sort_bubble)
                current_algorithm = line[:-1]
            elif line and current_algorithm:
                # Extraire ID et résultat attendu (ex : 184898 BubbleSort)
                parts = line.split()
                function_id = int(parts[0])
                expected_result = parts[1]
                expected_results[function_id] = expected_result
    return expected_results

# Initialiser le modèle et la base de données
llm = OllamaLLM(model="llama3.1:latest", use_cache=False)
conn = sqlite3.connect('data/algorithms.db')
cursor = conn.cursor()

# Charger les résultats attendus
expected_results = parse_expected_results("functions.txt")

# Demander le nom de la feuille à l'utilisateur
sheet_name = input("Entrez le nom de la feuille où insérer les données : ")

# Prompt utilisé pour Ollama
prompt_template = (
    "You will see one function written in C that tries to implement an algorithm. "
    "If the algorithm corresponds to the function's name, respond only with the algorithm name (for example, 'Bubble Sort') and nothing else. "
    "If other functions are being defined, please respond only with 'other' (and nothing else ! Please !). "
    "If the code implemented by the function doesn't correspond to the function's name (for example the function's name is 'tri_bulle' (Bubble Sort) but it implements Bogo Sort), please answer only and only with 'other'. "
    "This means that I don't want to see you responding with anything else than either 'other' or the generic english name of the algorithm if it is consistent with the name of the function."

    # "You will see one and only one function written in C that tries to implement an algorithm. "
    # "If the algorithm corresponds to the function's name, respond only with the algorithm name (for example, 'Bogo Sort') and nothing else. Watch closely because there could be tiny mistakes in the code which could alter its functionning. "
    # "If you see other functions being defined after the first one, please respond with 'other' (and nothing else ! Please !) altogether. "
    # "If the code implemented by the function doesn't match with the function's name (for example the function's name is 'tri_bulle' (Bubble Sort) but it implements Quick Sort), please answer only and only with 'other'. "
    # "This means that I don't want to see you responding with anything else than the bare minimum, either 'other' or the generic English name of the algorithm if it is consistent with the name of the function. "
)

# Préparer le fichier Excel
file_path = "algorithm_results.xlsx"
if os.path.exists(file_path):
    workbook = openpyxl.load_workbook(file_path)
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(sheet_name)
else:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name

# Vérifier si le prompt est déjà ajouté
if not sheet.cell(row=1, column=1).value:
    sheet.cell(row=1, column=1, value="Prompt utilisé :")
    sheet.cell(row=1, column=2, value=prompt_template)

# Trouver la prochaine colonne libre pour les nouvelles réponses
column_index = len(sheet[2]) + 1
column_letter = openpyxl.utils.get_column_letter(column_index)
sheet.cell(row=2, column=column_index, value=f"Model Response {column_index - 2}")

# Définir le style pour les cellules contenant "autre"
blue_font = Font(color="0000FF")  # Bleu

# Mesurer le temps de départ
start_time = time.time()

# Itérer sur chaque identifiant de fonction
for function_id, expected_result in expected_results.items():
    # Récupérer le code de la fonction à partir de la base de données
    cursor.execute(
        "SELECT * FROM algorithms WHERE id = ?",
        (function_id,)
    )
    function_instance = cursor.fetchone()

    if function_instance:
        algorithm_code = function_instance[3]  # Le code de la fonction

        # Construire le prompt pour Ollama
        prompt = prompt_template + f"{algorithm_code} "

        # Exécuter le modèle
        response = llm.invoke(prompt)
        print(f"Function ID: {function_id}, Expected: {expected_result}, Model Response: {response}")

        # Vérifier si la ligne existe déjà pour cet ID
        for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=1, max_col=1):
            if row[0].value == function_id:
                cell = sheet.cell(row=row[0].row, column=column_index, value=response)
                if response.lower().startswith("other"):
                    cell.font = blue_font
                break
        else:
            # Ajouter une nouvelle ligne si l'ID n'est pas trouvé
            new_row = [function_id, expected_result] + [""] * (column_index - 3) + [response]
            sheet.append(new_row)
            # Appliquer le style si nécessaire
            if response.lower().startswith("autre"):
                sheet.cell(row=sheet.max_row, column=column_index).font = blue_font

# Mesurer le temps d'arrivée
end_time = time.time()

# Calculer et afficher le temps écoulé
elapsed_time = end_time - start_time
print(f"Temps d'exécution de la boucle : {elapsed_time:.2f} secondes")

# Enregistrer le fichier Excel
workbook.save(file_path)

# Afficher où le fichier Excel a été sauvegardé
print(f"Fichier Excel sauvegardé à : {file_path}")

# Fermer la base de données
conn.close()
